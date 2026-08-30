from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path, PurePosixPath
import re
from typing import Literal
from uuid import uuid5
import zipfile

from openpyxl import load_workbook  # type: ignore[import-untyped]

from due_diligence_agent.application.services.table_normalization_service import (
    TableNormalizationService,
)
from due_diligence_agent.domain.artifacts.models import Artifact, SourceLocator
from due_diligence_agent.domain.documents.tabular import (
    NormalizedCell,
    NormalizedTable,
    SheetVisibility,
    SpreadsheetParseResult,
)
from due_diligence_agent.domain.evidence.models import EvidenceFact
from due_diligence_agent.domain.metrics.definitions import PUBLIC_METRIC_DEFINITIONS
from due_diligence_agent.domain.metrics.startup import STARTUP_METRIC_DEFINITIONS
from due_diligence_agent.ports.repositories import ArtifactStore


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLSM_MEDIA_TYPE = "application/vnd.ms-excel.sheet.macroEnabled.12"
CSV_MEDIA_TYPES = {"text/csv", "application/csv", "text/plain"}
_PERIOD_YEAR = re.compile(r"^\d{4}$")
_PERIOD_QUARTER = re.compile(
    r"^(?:(\d{4})\s*[-/]?\s*Q([1-4])|Q([1-4])\s*[-/]?\s*(\d{4}))$",
    re.IGNORECASE,
)
_PLAIN_NUMBER = re.compile(r"^\d+(?:\.\d+)?$")
_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_PUBLIC_FACT_NAMES = frozenset(
    slot.fact_name for definition in PUBLIC_METRIC_DEFINITIONS for slot in definition.slots
)
_STARTUP_FACT_NAMES = frozenset(
    slot.fact_name for definition in STARTUP_METRIC_DEFINITIONS for slot in definition.slots
)
_CANONICAL_FACT_NAMES = _PUBLIC_FACT_NAMES | _STARTUP_FACT_NAMES
_MARKET_AS_OF_FACT_NAMES = frozenset(
    slot.fact_name
    for definition in PUBLIC_METRIC_DEFINITIONS
    for slot in definition.slots
    if slot.period_role == "market_as_of"
)
_FACT_LABEL_ALIASES = {
    **{name: name for name in _CANONICAL_FACT_NAMES},
    **{name.replace("_", " "): name for name in _CANONICAL_FACT_NAMES},
    "capex": "capital_expenditures",
    "cash & equivalents": "cash_and_equivalents",
    "cash and cash equivalents": "cash_and_equivalents",
    "cash flow from operations": "operating_cash_flow",
    "diluted weighted average shares": "weighted_average_diluted_shares",
    "total revenues": "revenue",
}


@dataclass(frozen=True, slots=True)
class SpreadsheetLimits:
    max_bytes: int = 25 * 1024 * 1024
    max_rows: int = 50_000
    max_columns: int = 512
    max_field_chars: int = 100_000
    max_xlsx_entries: int = 10_000
    max_xlsx_entry_uncompressed_bytes: int = 100 * 1024 * 1024
    max_xlsx_total_uncompressed_bytes: int = 500 * 1024 * 1024
    max_xlsx_compression_ratio: int = 200

    def __post_init__(self) -> None:
        configured = (
            self.max_bytes,
            self.max_rows,
            self.max_columns,
            self.max_field_chars,
            self.max_xlsx_entries,
            self.max_xlsx_entry_uncompressed_bytes,
            self.max_xlsx_total_uncompressed_bytes,
            self.max_xlsx_compression_ratio,
        )
        if min(configured) < 1:
            raise ValueError("spreadsheet limits must be positive")


class _SpreadsheetQuotaError(RuntimeError):
    def __init__(self, error_code: str) -> None:
        super().__init__(error_code)
        self.error_code = error_code


class SpreadsheetParser:
    parser_name = "openpyxl+csv"
    parser_version = "1"

    def __init__(
        self,
        *,
        artifact_store: ArtifactStore,
        normalizer: TableNormalizationService,
        limits: SpreadsheetLimits | None = None,
    ) -> None:
        self._artifact_store = artifact_store
        self._normalizer = normalizer
        self._limits = limits or SpreadsheetLimits()

    def parse(self, artifact: Artifact, *, no_network: bool = True) -> SpreadsheetParseResult:
        del no_network  # This adapter has no network-capable dependency or code path.
        try:
            payload = self._artifact_store.read_bytes(artifact.content_hash)
        except Exception:
            return SpreadsheetParseResult.outcome(
                artifact_id=artifact.id,
                status="malformed",
                error_code="spreadsheet_storage_error",
            )
        if len(payload) > self._limits.max_bytes:
            return SpreadsheetParseResult.outcome(
                artifact_id=artifact.id,
                status="quota_exceeded",
                error_code="spreadsheet_byte_limit_exceeded",
            )

        source_suffix = Path(artifact.source).suffix.casefold()
        if source_suffix == ".xlsm" or artifact.mime_type.casefold() == XLSM_MEDIA_TYPE.casefold():
            return SpreadsheetParseResult.outcome(
                artifact_id=artifact.id,
                status="rejected",
                error_code="macro_enabled_workbook_rejected",
            )
        if payload.startswith(b"PK\x03\x04"):
            return self._parse_xlsx(artifact, payload)
        if source_suffix == ".csv" or artifact.mime_type.casefold() in CSV_MEDIA_TYPES:
            return self._parse_csv(artifact, payload)
        return SpreadsheetParseResult.outcome(
            artifact_id=artifact.id,
            status="unsupported",
            error_code="unsupported_spreadsheet_type",
        )

    def _parse_xlsx(self, artifact: Artifact, payload: bytes) -> SpreadsheetParseResult:
        try:
            _preflight_xlsx(payload, self._limits)
        except _SpreadsheetQuotaError as exc:
            return SpreadsheetParseResult.outcome(
                artifact_id=artifact.id,
                status="quota_exceeded",
                error_code=exc.error_code,
            )
        except _UnsafeSpreadsheetError as exc:
            return SpreadsheetParseResult.outcome(
                artifact_id=artifact.id,
                status="rejected",
                error_code=exc.error_code,
            )
        rejected_code = _unsafe_ooxml_code(payload)
        if rejected_code is not None:
            return SpreadsheetParseResult.outcome(
                artifact_id=artifact.id,
                status="rejected",
                error_code=rejected_code,
            )
        try:
            values_book = load_workbook(
                BytesIO(payload),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
            formulas_book = load_workbook(
                BytesIO(payload),
                read_only=True,
                data_only=False,
                keep_links=False,
            )
            try:
                tables = self._xlsx_tables(artifact, values_book, formulas_book)
            finally:
                values_book.close()
                formulas_book.close()
        except _SpreadsheetQuotaError as exc:
            return SpreadsheetParseResult.outcome(
                artifact_id=artifact.id,
                status="quota_exceeded",
                error_code=exc.error_code,
            )
        except Exception:
            return SpreadsheetParseResult.outcome(
                artifact_id=artifact.id,
                status="malformed",
                error_code="malformed_xlsx",
            )
        facts = _evidence_facts(artifact, tables)
        partial = any(
            cell.formula_cached is False for table in tables for cell in table.cells
        )
        return SpreadsheetParseResult(
            artifact_id=artifact.id,
            tables=tables,
            evidence_facts=facts,
            status="partial" if partial else "parsed",
            error_code="formula_cached_value_unavailable" if partial else None,
        )

    def _xlsx_tables(
        self,
        artifact: Artifact,
        values_book: object,
        formulas_book: object,
    ) -> list[NormalizedTable]:
        tables: list[NormalizedTable] = []
        for sheet_name in values_book.sheetnames:  # type: ignore[attr-defined]
            values_sheet = values_book[sheet_name]  # type: ignore[index]
            formulas_sheet = formulas_book[sheet_name]  # type: ignore[index]
            raw_cells, row_count, column_count = self._xlsx_cells(values_sheet, formulas_sheet)
            cells = _normalize_raw_cells(
                artifact=artifact,
                table_name=sheet_name,
                raw_cells=raw_cells,
                locator_kind="xlsx_cell",
            )
            visibility = _visibility(values_sheet.sheet_state)
            tables.append(
                self._normalizer.normalize_and_store(
                    artifact=artifact,
                    name=sheet_name,
                    cells=cells,
                    visibility=visibility,
                    row_count=row_count,
                    column_count=column_count,
                )
            )
        return tables

    def _xlsx_cells(
        self,
        values_sheet: object,
        formulas_sheet: object,
    ) -> tuple[list[_RawCell], int, int]:
        formulas: dict[tuple[int, int], str] = {}
        for row_index, row in enumerate(formulas_sheet.iter_rows(), start=1):  # type: ignore[attr-defined]
            if row_index > self._limits.max_rows:
                raise _SpreadsheetQuotaError("xlsx_row_limit_exceeded")
            if len(row) > self._limits.max_columns:
                raise _SpreadsheetQuotaError("xlsx_column_limit_exceeded")
            for column_index, cell in enumerate(row, start=1):
                if cell.data_type == "f":
                    formulas[(row_index, column_index)] = str(cell.value)

        raw_cells: list[_RawCell] = []
        row_count = 0
        column_count = 0
        for row_index, row in enumerate(values_sheet.iter_rows(), start=1):  # type: ignore[attr-defined]
            if row_index > self._limits.max_rows:
                raise _SpreadsheetQuotaError("xlsx_row_limit_exceeded")
            if len(row) > self._limits.max_columns:
                raise _SpreadsheetQuotaError("xlsx_column_limit_exceeded")
            row_count = row_index
            column_count = max(column_count, len(row))
            for column_index, cell in enumerate(row, start=1):
                formula = formulas.get((row_index, column_index))
                if cell.value is None and formula is None:
                    continue
                raw_cells.append(
                    _RawCell(
                        row=row_index,
                        column=column_index,
                        coordinate=cell.coordinate,
                        value=_typed_xlsx_value(cell.value),
                        number_format=str(cell.number_format or ""),
                        formula=formula,
                    )
                )
        for (row_index, column_index), formula in formulas.items():
            if any(
                cell.row == row_index and cell.column == column_index for cell in raw_cells
            ):
                continue
            raw_cells.append(
                _RawCell(
                    row=row_index,
                    column=column_index,
                    coordinate=_coordinate(column_index, row_index),
                    value=None,
                    number_format="",
                    formula=formula,
                )
            )
            row_count = max(row_count, row_index)
            column_count = max(column_count, column_index)
        return raw_cells, row_count, column_count

    def _parse_csv(self, artifact: Artifact, payload: bytes) -> SpreadsheetParseResult:
        decoded = _decode_csv(payload)
        if decoded is None:
            return SpreadsheetParseResult.outcome(
                artifact_id=artifact.id,
                status="malformed",
                error_code="csv_encoding_not_allowed",
            )
        text, encoding = decoded
        previous_limit = csv.field_size_limit()
        try:
            csv.field_size_limit(self._limits.max_field_chars)
            reader = csv.reader(StringIO(text, newline=""), strict=True)
            rows: list[list[str]] = []
            for row_index, row in enumerate(reader, start=1):
                if row_index > self._limits.max_rows:
                    raise _SpreadsheetQuotaError("csv_row_limit_exceeded")
                if len(row) > self._limits.max_columns:
                    raise _SpreadsheetQuotaError("csv_column_limit_exceeded")
                rows.append(row)
        except _SpreadsheetQuotaError as exc:
            return SpreadsheetParseResult.outcome(
                artifact_id=artifact.id,
                status="quota_exceeded",
                error_code=exc.error_code,
            )
        except csv.Error as exc:
            code = (
                "csv_field_limit_exceeded"
                if "field larger than field limit" in str(exc).casefold()
                else "malformed_csv"
            )
            status: Literal["malformed", "quota_exceeded"] = (
                "quota_exceeded" if code == "csv_field_limit_exceeded" else "malformed"
            )
            return SpreadsheetParseResult.outcome(
                artifact_id=artifact.id,
                status=status,
                error_code=code,
            )
        finally:
            csv.field_size_limit(previous_limit)

        raw_cells = [
            _RawCell(
                row=row_index,
                column=column_index,
                coordinate=f"R{row_index}C{column_index}",
                value=_typed_csv_value(value),
                number_format="",
                formula=None,
            )
            for row_index, row in enumerate(rows, start=1)
            for column_index, value in enumerate(row, start=1)
            if value != ""
        ]
        name = Path(artifact.source).stem
        cells = _normalize_raw_cells(
            artifact=artifact,
            table_name=name,
            raw_cells=raw_cells,
            locator_kind="csv_cell",
        )
        table = self._normalizer.normalize_and_store(
            artifact=artifact,
            name=name,
            cells=cells,
            visibility="visible",
            row_count=len(rows),
            column_count=max((len(row) for row in rows), default=0),
        )
        return SpreadsheetParseResult(
            artifact_id=artifact.id,
            tables=[table],
            evidence_facts=_evidence_facts(artifact, [table]),
            status="parsed",
            encoding=encoding,
        )


@dataclass(frozen=True, slots=True)
class _RawCell:
    row: int
    column: int
    coordinate: str
    value: bool | date | Decimal | str | None
    number_format: str
    formula: str | None


class _UnsafeSpreadsheetError(RuntimeError):
    def __init__(self, error_code: str) -> None:
        super().__init__(error_code)
        self.error_code = error_code


def _normalize_raw_cells(
    *,
    artifact: Artifact,
    table_name: str,
    raw_cells: list[_RawCell],
    locator_kind: Literal["xlsx_cell", "csv_cell"],
) -> list[NormalizedCell]:
    by_position = {(cell.row, cell.column): cell for cell in raw_cells}
    cells: list[NormalizedCell] = []
    for raw in sorted(raw_cells, key=lambda cell: (cell.row, cell.column)):
        label = _label_hint(raw, by_position)
        canonical_fact_name = _canonical_fact_name(label)
        period = _period_hint(raw, by_position)
        unit = _unit_hint(raw, label, by_position)
        numeric = isinstance(raw.value, Decimal)
        status: Literal["candidate", "insufficient_data"] = (
            "candidate"
            if numeric
            and canonical_fact_name
            and _is_accepted_spreadsheet_fact_period(canonical_fact_name, period)
            and unit
            else "insufficient_data"
        )
        locator_value = (
            _xlsx_locator_value(table_name, raw.coordinate)
            if locator_kind == "xlsx_cell"
            else raw.coordinate
        )
        cells.append(
            NormalizedCell(
                row=raw.row,
                column=raw.column,
                label=label,
                period=period,
                value=raw.value,
                unit=unit,
                locator=SourceLocator(
                    kind=locator_kind,
                    value=locator_value,
                    artifact_id=artifact.id,
                    table=table_name,
                    cell=raw.coordinate,
                ),
                status=status,
                formula_cached=(raw.value is not None) if raw.formula is not None else None,
            )
        )
    return cells


def _label_hint(raw: _RawCell, cells: dict[tuple[int, int], _RawCell]) -> str | None:
    for column in range(raw.column - 1, 0, -1):
        candidate = cells.get((raw.row, column))
        if candidate is None or not isinstance(candidate.value, str):
            continue
        normalized = candidate.value.strip()
        if normalized and _normalize_period(normalized) is None:
            return normalized
    return None


def _period_hint(raw: _RawCell, cells: dict[tuple[int, int], _RawCell]) -> str | None:
    for row in range(raw.row - 1, 0, -1):
        candidate = cells.get((row, raw.column))
        if candidate is None:
            continue
        if isinstance(candidate.value, date):
            return str(candidate.value.year)
        if isinstance(candidate.value, str):
            period = _normalize_period(candidate.value)
            if period is not None:
                return period
    return None


def _unit_hint(
    raw: _RawCell,
    label: str | None,
    cells: dict[tuple[int, int], _RawCell],
) -> str | None:
    number_format = raw.number_format.casefold()
    if "%" in number_format:
        return "percent"
    if "$" in number_format:
        return "USD"
    if "€" in number_format:
        return "EUR"
    if "£" in number_format:
        return "GBP"
    contexts = [label or ""]
    for row in range(raw.row - 1, 0, -1):
        candidate = cells.get((row, raw.column))
        if candidate is not None and isinstance(candidate.value, str):
            contexts.append(candidate.value)
    context = " ".join(contexts)
    if "%" in context or re.search(r"\bpercent(?:age)?\b", context, re.IGNORECASE):
        return "percent"
    for code in ("USD", "EUR", "GBP"):
        if re.search(rf"\b{code}\b", context, re.IGNORECASE):
            multiplier = re.search(r"\b(thousand|million|billion)s?\b", context, re.IGNORECASE)
            return f"{code} {multiplier.group(1).casefold()}s" if multiplier else code
    return None


def _normalize_period(value: str) -> str | None:
    normalized = value.strip()
    if _PERIOD_YEAR.fullmatch(normalized):
        return normalized
    quarter = _PERIOD_QUARTER.fullmatch(normalized)
    if quarter:
        year = quarter.group(1) or quarter.group(4)
        quarter_number = quarter.group(2) or quarter.group(3)
        return f"{year}-Q{quarter_number}"
    return None


def _typed_xlsx_value(value: object) -> bool | date | Decimal | str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    if isinstance(value, int | float):
        converted = Decimal(str(value))
        return converted if converted.is_finite() else None
    return str(value)


def _typed_csv_value(value: str) -> bool | date | Decimal | str:
    stripped = value.strip()
    if not stripped:
        return value
    if stripped[0] in "=+-@":
        return value
    if stripped.casefold() == "true":
        return True
    if stripped.casefold() == "false":
        return False
    if _ISO_DATE.fullmatch(stripped):
        try:
            return date.fromisoformat(stripped)
        except ValueError:
            return value
    numeric = stripped.replace(",", "")
    if _PLAIN_NUMBER.fullmatch(numeric):
        try:
            converted = Decimal(numeric)
            return converted if converted.is_finite() else value
        except InvalidOperation:
            return value
    return value


def _unsafe_ooxml_code(payload: bytes) -> str | None:
    try:
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            names = [name.casefold() for name in archive.namelist()]
            if any("vbaproject" in name or name.endswith(".bin") for name in names):
                return "macro_enabled_workbook_rejected"
            if any(name.startswith("xl/externallinks/") for name in names):
                return "external_links_rejected"
            for name in archive.namelist():
                if not name.casefold().endswith(".rels"):
                    continue
                relationship_data = archive.read(name)
                if b"TargetMode=\"External\"" in relationship_data or b"TargetMode='External'" in relationship_data:
                    return "external_links_rejected"
    except (OSError, zipfile.BadZipFile, KeyError):
        return None
    return None


def _preflight_xlsx(payload: bytes, limits: SpreadsheetLimits) -> None:
    try:
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            entries = archive.infolist()
    except (OSError, zipfile.BadZipFile):
        raise _UnsafeSpreadsheetError("malformed_xlsx_archive") from None
    if len(entries) > limits.max_xlsx_entries:
        raise _SpreadsheetQuotaError("xlsx_archive_entry_limit_exceeded")
    total_size = 0
    for entry in entries:
        if _unsafe_archive_member_name(entry.filename):
            raise _UnsafeSpreadsheetError("xlsx_archive_member_invalid")
        if entry.file_size > limits.max_xlsx_entry_uncompressed_bytes:
            raise _SpreadsheetQuotaError("xlsx_archive_entry_size_exceeded")
        total_size += entry.file_size
        if total_size > limits.max_xlsx_total_uncompressed_bytes:
            raise _SpreadsheetQuotaError("xlsx_archive_total_size_exceeded")
        if entry.file_size and entry.file_size / max(entry.compress_size, 1) > (
            limits.max_xlsx_compression_ratio
        ):
            raise _SpreadsheetQuotaError("xlsx_archive_compression_ratio_exceeded")


def _unsafe_archive_member_name(value: str) -> bool:
    normalized = value.replace("\\", "/")
    path = PurePosixPath(normalized)
    return (
        not normalized
        or "\x00" in normalized
        or normalized.startswith(("/", "//"))
        or bool(re.match(r"^[A-Za-z]:", normalized))
        or ".." in path.parts
    )


def _decode_csv(payload: bytes) -> tuple[str, str] | None:
    candidates: list[tuple[str, str]]
    if payload.startswith(b"\xef\xbb\xbf"):
        candidates = [("utf-8-sig", "utf-8-sig")]
    elif payload.startswith(b"\xff\xfe"):
        candidates = [("utf-16", "utf-16-le")]
    elif payload.startswith(b"\xfe\xff"):
        candidates = [("utf-16", "utf-16-be")]
    else:
        candidates = [("utf-8", "utf-8"), ("cp1252", "cp1252")]
    for codec, label in candidates:
        try:
            return payload.decode(codec), label
        except UnicodeDecodeError:
            continue
    return None


def _evidence_facts(
    artifact: Artifact,
    tables: list[NormalizedTable],
) -> list[EvidenceFact]:
    facts: list[EvidenceFact] = []
    for table in tables:
        for cell in table.cells:
            canonical_fact_name = _canonical_fact_name(cell.label)
            period = cell.period
            if (
                cell.status != "candidate"
                or not isinstance(cell.value, Decimal)
                or canonical_fact_name is None
                or period is None
                or not _is_accepted_spreadsheet_fact_period(canonical_fact_name, period)
                or cell.unit is None
            ):
                continue
            facts.append(
                EvidenceFact(
                    id=uuid5(
                        artifact.id,
                        "\x1f".join(
                            (
                                cell.locator.value,
                                canonical_fact_name,
                                period,
                                cell.unit,
                            )
                        ),
                    ),
                    artifact_id=artifact.id,
                    name=canonical_fact_name,
                    value=cell.value,
                    value_type="decimal",
                    unit=cell.unit,
                    period=period,
                    locator=cell.locator,
                    sensitivity=artifact.sensitivity,
                    confidence=Decimal("0.80"),
                    extraction_method="spreadsheet_parser",
                    supporting_text_hash=table.snapshot_hash,
                    metadata={"table": table.name},
                )
            )
    return facts


def _canonical_fact_name(label: str | None) -> str | None:
    if label is None:
        return None
    normalized = re.sub(r"\s+", " ", label).strip().casefold()
    return _FACT_LABEL_ALIASES.get(normalized)


def _is_accepted_spreadsheet_fact_period(
    canonical_fact_name: str,
    period: str | None,
) -> bool:
    if canonical_fact_name in _MARKET_AS_OF_FACT_NAMES:
        return False
    return bool(period and re.fullmatch(r"(?:\d{4}|\d{4}-Q[1-4])", period))


def _xlsx_locator_value(sheet_name: str, coordinate: str) -> str:
    if "!" not in sheet_name and "'" not in sheet_name:
        return f"{sheet_name}!{coordinate}"
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'!{coordinate}"


def _visibility(value: str) -> SheetVisibility:
    if value == "hidden":
        return "hidden"
    if value == "veryHidden":
        return "veryHidden"
    return "visible"


def _coordinate(column: int, row: int) -> str:
    letters = ""
    current = column
    while current:
        current, remainder = divmod(current - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{row}"
