from __future__ import annotations

import json
import re
import zipfile
from hashlib import sha256
from pathlib import Path
from typing import Any, cast

from docx import Document
from openpyxl import load_workbook  # type: ignore[import-untyped]
from PIL import Image
from pypdf import PdfReader


FIXTURE_ROOT = Path("tests/fixtures/startup_synthetic_v1")
EXPECTED_FORMATS = {"pdf", "docx", "png", "jpeg", "csv", "xlsx", "safe_zip"}
PRIVACY_SENTINELS = re.compile(
    r"(?i)(sk-|api[_-]?key|access[_-]?token|authorization|bearer|password|secret|[a-z]:[\\/]|file://|\\\\|[\w.+-]+@[\w.-]+\.[a-z]{2,})"
)


def test_startup_synthetic_manifest_hashes_privacy_and_contract_hash_are_stable() -> None:
    first = _load_and_validate_contract()
    second = _load_and_validate_contract()

    assert first == second
    assert first["contract_hash"] == "sha256:" + _hash_json(first["expected_contracts"])


def test_startup_synthetic_matrix_covers_required_models_and_contracts() -> None:
    contract = _load_and_validate_contract()
    cases = cast(dict[str, Any], contract["expected_contracts"]["cases"])

    assert tuple(sorted(cases)) == ("marketplace", "pre_revenue_service", "saas", "transactional")
    assert contract["manifest"]["active_format_matrix"] == sorted(EXPECTED_FORMATS)
    assert {
        item
        for case in cast(dict[str, Any], contract["manifest"]["cases"]).values()
        for item in case["formats"]
    } == EXPECTED_FORMATS

    for case_name, case in cases.items():
        profile_fields = cast(dict[str, Any], case["profile_fields"])
        assert {
            "problem",
            "solution",
            "customer_segment",
            "business_model",
            "traction",
            "market",
            "competition",
            "financials",
        } <= set(profile_fields)
        assert all("status" in field for field in profile_fields.values())
        assert all("expected_evidence_status" in field for field in profile_fields.values())
        assert case["expected_contradictions"], case_name
        assert case["expected_unsupported_claims"], case_name
        assert case["expected_gap_codes"], case_name
        assert case["market_research"]["source_mode"] == "frozen"
        assert case["market_research"]["as_of"] == contract["manifest"]["as_of"]
        assert case["market_research"]["source_ids"]
        assert case["readiness"]["status"] in {"ready_for_first_sales", "needs_validation", "not_ready"}
        assert len(case["readiness"]["blocking_questions"]) <= 3
        assert set(case["trace_to_report_lineage"]) == {
            "input_documents",
            "profile_fields",
            "metrics",
            "readiness",
            "research_sources",
            "competitors",
            "market_sizing",
            "risks",
            "report_sections",
        }

    assert cases["saas"]["market_research"]["market_sizing_status"] == "calculated"
    assert cases["pre_revenue_service"]["market_research"]["market_sizing_status"] == "insufficient_data"
    assert "market_research" in contract["expected_contracts"]["report_sections"]
    assert contract["expected_contracts"]["trace_contract"]["requires_report_lineage"] is True


def _load_and_validate_contract() -> dict[str, Any]:
    manifest = _read_json(FIXTURE_ROOT / "manifest.json")
    assert manifest["schema_version"] == "startup_synthetic_fixture_manifest@1"
    assert manifest["network_policy"] == "no_external_network"
    assert manifest["privacy_policy"] == "synthetic_no_secrets_no_emails_no_local_paths"

    total_bytes = 0
    expected_entry = cast(dict[str, Any], manifest["expected_contracts"])
    expected_contracts = _read_verified_json(expected_entry)
    total_bytes += int(expected_entry["bytes"])

    assert manifest["active_format_matrix"] == sorted(EXPECTED_FORMATS)
    for case_name, case in cast(dict[str, Any], manifest["cases"]).items():
        assert case_name in expected_contracts["cases"]
        assert set(case["formats"]) <= EXPECTED_FORMATS
        assert set(case["formats"])
        for file_entry in cast(dict[str, Any], case["files"]).values():
            total_bytes += int(file_entry["bytes"])
            _verify_file_entry(file_entry)
            assert file_entry["format"] in EXPECTED_FORMATS
            _verify_supported_format_signature(file_entry)

    assert total_bytes <= int(manifest["max_total_fixture_bytes"])
    return {
        "manifest": manifest,
        "expected_contracts": expected_contracts,
        "contract_hash": "sha256:" + _hash_json(expected_contracts),
    }


def _read_verified_json(entry: dict[str, Any]) -> dict[str, Any]:
    _verify_file_entry(entry)
    return _read_json(FIXTURE_ROOT / str(entry["path"]))


def _verify_file_entry(entry: dict[str, Any]) -> None:
    relative_path = str(entry["path"]) if "path" in entry else _find_manifest_path(entry)
    path = FIXTURE_ROOT / relative_path
    assert path.is_file()
    payload = path.read_bytes()
    assert len(payload) == int(entry["bytes"])
    assert sha256(payload).hexdigest() == entry["sha256"]
    try:
        text = payload.decode("utf-8")
    except UnicodeDecodeError:
        text = ""
    assert PRIVACY_SENTINELS.search(text) is None
    for member in entry.get("archive_members", ()):
        member_path = str(member["path"])
        assert "\\" not in member_path
        assert not member_path.startswith("/")
        assert ".." not in Path(member_path).parts
        assert int(member["bytes"]) <= int(entry["max_member_bytes"])


def _verify_supported_format_signature(entry: dict[str, Any]) -> None:
    relative_path = str(entry["path"]) if "path" in entry else _find_manifest_path(entry)
    path = FIXTURE_ROOT / relative_path
    payload = path.read_bytes()
    signatures = {
        "pdf": b"%PDF-",
        "docx": b"PK",
        "png": b"\x89PNG",
        "jpeg": b"\xff\xd8\xff",
        "xlsx": b"PK",
        "safe_zip": b"PK",
    }
    expected = signatures.get(str(entry["format"]))
    if expected is not None:
        assert payload.startswith(expected)
    match str(entry["format"]):
        case "pdf":
            reader = PdfReader(path)
            assert len(reader.pages) >= 1
            assert "\n".join(page.extract_text() or "" for page in reader.pages).strip()
        case "docx":
            document = Document(str(path))
            assert "\n".join(paragraph.text for paragraph in document.paragraphs).strip()
        case "xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                values = [
                    cell
                    for sheet in workbook.worksheets
                    for row in sheet.iter_rows(values_only=True)
                    for cell in row
                    if cell not in (None, "")
                ]
                assert values
            finally:
                workbook.close()
        case "png" | "jpeg":
            with Image.open(path) as image:
                image.verify()
                assert image.format in {"PNG", "JPEG"}
        case "safe_zip":
            _verify_safe_zip_members(path, entry)


def _verify_safe_zip_members(path: Path, entry: dict[str, Any]) -> None:
    expected_members = {
        str(member["path"]): member for member in cast(list[dict[str, Any]], entry["archive_members"])
    }
    with zipfile.ZipFile(path) as archive:
        infos = archive.infolist()
        assert {info.filename for info in infos} == set(expected_members)
        for info in infos:
            member = expected_members[info.filename]
            payload = archive.read(info.filename)
            assert info.file_size == int(member["bytes"])
            assert len(payload) == int(member["bytes"])
            assert sha256(payload).hexdigest() == member["sha256"]
            assert payload.strip()


def _find_manifest_path(entry: dict[str, Any]) -> str:
    manifest = _read_json(FIXTURE_ROOT / "manifest.json")
    for case in cast(dict[str, Any], manifest["cases"]).values():
        for relative_path, candidate in cast(dict[str, Any], case["files"]).items():
            if candidate == entry:
                return str(relative_path)
    raise AssertionError("manifest file entry not found")


def _read_json(path: Path) -> dict[str, Any]:
    return cast(dict[str, Any], json.loads(path.read_text(encoding="utf-8")))


def _hash_json(value: object) -> str:
    return sha256(json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True).encode("utf-8")).hexdigest()
