from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
import json
from pathlib import Path
from uuid import UUID, uuid4
import zipfile

import duckdb
from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel
import pytest

import due_diligence_agent.adapters.documents.spreadsheet_parser as spreadsheet_module
from due_diligence_agent.adapters.documents.spreadsheet_parser import (
    SpreadsheetLimits,
    SpreadsheetParser,
)
from due_diligence_agent.adapters.local_storage.artifact_store import LocalArtifactStore
from due_diligence_agent.application.services.table_normalization_service import (
    TableNormalizationService,
)
from due_diligence_agent.domain.artifacts.models import Artifact
from due_diligence_agent.domain.common import SensitivityClass


CASE_ID = UUID("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa")


@pytest.fixture
def spreadsheet_runtime(
    tmp_path: Path,
) -> tuple[SpreadsheetParser, LocalArtifactStore, Path]:
    store = LocalArtifactStore(tmp_path / "artifact-store")
    database_path = tmp_path / "normalized.duckdb"
    normalizer = TableNormalizationService(artifact_store=store, database_path=database_path)
    return SpreadsheetParser(artifact_store=store, normalizer=normalizer), store, database_path


def test_xlsx_value_preserves_sheet_and_cell_locator_with_duplicate_labels(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "P&L"
    sheet["C11"] = "2025"
    sheet["A12"] = "Revenue"
    sheet["C12"] = 1_800_000
    sheet["C12"].number_format = '$#,##0'
    sheet["A13"] = "Revenue"
    sheet["C13"] = 1_900_000
    sheet["C13"].number_format = '$#,##0'
    artifact = _persist(store, _xlsx_bytes(workbook), "financial-model.xlsx")

    table = parser.parse(artifact, no_network=True).tables[0]
    revenue = table.find(label="Revenue", period="2025")

    assert revenue is not None
    assert revenue.locator.kind == "xlsx_cell"
    assert revenue.locator.value == "P&L!C12"
    assert revenue.value == Decimal("1800000")
    assert [cell.locator.value for cell in table.cells if cell.label == "Revenue"] == [
        "P&L!C12",
        "P&L!C13",
    ]


def test_xlsx_locator_is_stable_for_sheet_names_with_spaces_and_quotes(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Founder Plan's"
    sheet["A1"] = "2026"
    sheet["A2"] = "ARR"
    sheet["B2"] = 250
    sheet["B2"].number_format = '[$$-409]#,##0'
    artifact = _persist(store, _xlsx_bytes(workbook), "quoted-sheet.xlsx")

    table = parser.parse(artifact, no_network=True).tables[0]

    assert next(cell for cell in table.cells if cell.value == Decimal("250")).locator.value == (
        "'Founder Plan''s'!B2"
    )


def test_xlsx_locator_escapes_separator_and_embedded_quote_unambiguously(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ops!'North"
    sheet["A1"] = "2026"
    sheet["A2"] = "Revenue"
    sheet["B2"] = 250
    sheet["B2"].number_format = '$#,##0'
    artifact = _persist(store, _xlsx_bytes(workbook), "escaped-sheet.xlsx")

    cell = next(
        cell
        for cell in parser.parse(artifact, no_network=True).tables[0].cells
        if cell.value == Decimal("250")
    )

    assert cell.locator.value == "'Ops!''North'!B2"
    assert cell.locator.table == "Ops!'North"
    assert cell.locator.cell == "B2"


def test_formula_without_cached_value_is_not_invented(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Forecast"
    sheet["A1"] = "=1+1"
    artifact = _persist(store, _xlsx_bytes(workbook), "formula-only.xlsx")

    cell = parser.parse(artifact, no_network=True).tables[0].cells[0]

    assert cell.value is None
    assert cell.status == "insufficient_data"
    assert cell.formula_cached is False


@pytest.mark.parametrize(
    ("source", "media_type", "expected_code"),
    [
        (
            "model.xlsm",
            "application/vnd.ms-excel.sheet.macroEnabled.12",
            "macro_enabled_workbook_rejected",
        ),
        (
            "model.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "external_links_rejected",
        ),
    ],
)
def test_macros_and_external_links_are_rejected_with_typed_private_outcomes(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
    source: str,
    media_type: str,
    expected_code: str,
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    workbook = Workbook()
    workbook.active["A1"] = "SECRET-DO-NOT-LEAK"
    payload = _xlsx_bytes(workbook)
    if expected_code == "external_links_rejected":
        payload = _with_external_link(payload)
    artifact = _persist(store, payload, source, media_type=media_type)

    result = parser.parse(artifact, no_network=True)

    assert result.status == "rejected"
    assert result.error_code == expected_code
    assert "SECRET-DO-NOT-LEAK" not in repr(result)


@pytest.mark.parametrize(
    (
        "limit_overrides",
        "member_name",
        "member_payload",
        "expected_status",
        "expected_code",
    ),
    [
        (
            {},
            "../escape.xml",
            b"escape",
            "rejected",
            "xlsx_archive_member_invalid",
        ),
        (
            {"max_xlsx_entries": 1},
            None,
            b"",
            "quota_exceeded",
            "xlsx_archive_entry_limit_exceeded",
        ),
        (
            {"max_xlsx_entry_uncompressed_bytes": 1},
            None,
            b"",
            "quota_exceeded",
            "xlsx_archive_entry_size_exceeded",
        ),
        (
            {"max_xlsx_total_uncompressed_bytes": 1},
            None,
            b"",
            "quota_exceeded",
            "xlsx_archive_total_size_exceeded",
        ),
        (
            {"max_xlsx_compression_ratio": 2},
            "xl/media/bomb.bin",
            b"0" * 100_000,
            "quota_exceeded",
            "xlsx_archive_compression_ratio_exceeded",
        ),
    ],
    ids=["traversal", "entry-count", "entry-size", "total-size", "compression-ratio"],
)
def test_xlsx_archive_preflight_rejects_hostile_metadata_before_read_or_load(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    limit_overrides: dict[str, int],
    member_name: str | None,
    member_payload: bytes,
    expected_status: str,
    expected_code: str,
) -> None:
    workbook = Workbook()
    workbook.active["A1"] = "safe"
    payload = _xlsx_bytes(workbook)
    if member_name is not None:
        payload = _with_zip_member(payload, member_name, member_payload)
    store = LocalArtifactStore(tmp_path / "artifact-store")
    normalizer = TableNormalizationService(
        artifact_store=store,
        database_path=tmp_path / "normalized.duckdb",
    )
    parser = SpreadsheetParser(
        artifact_store=store,
        normalizer=normalizer,
        limits=SpreadsheetLimits(**limit_overrides),
    )
    artifact = _persist(store, payload, "hostile.xlsx")

    def forbidden_read(*_args: object, **_kwargs: object) -> bytes:
        pytest.fail("archive preflight must reject before member reads")

    def forbidden_load(*_args: object, **_kwargs: object) -> object:
        pytest.fail("archive preflight must reject before openpyxl load")

    monkeypatch.setattr(zipfile.ZipFile, "read", forbidden_read)
    monkeypatch.setattr(spreadsheet_module, "load_workbook", forbidden_load)

    result = parser.parse(artifact, no_network=True)

    assert result.status == expected_status
    assert result.error_code == expected_code


def test_xlsx_hidden_merged_and_typed_cells_are_deterministic(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    workbook = Workbook()
    visible = workbook.active
    visible.title = "Visible"
    visible.merge_cells("A1:B1")
    visible["A1"] = "Merged heading"
    visible["A2"] = "As of"
    visible["B2"] = to_excel(date(2024, 1, 1), workbook.epoch)
    visible["B2"].number_format = "yyyy-mm-dd"
    visible["A3"] = "Audited"
    visible["B3"] = True
    visible["A4"] = "Margin"
    visible["B4"] = 0.25
    visible["B4"].number_format = "0.0%"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "hidden-value"
    very_hidden = workbook.create_sheet("Very Hidden")
    very_hidden.sheet_state = "veryHidden"
    very_hidden["A1"] = "very-hidden-value"
    artifact = _persist(store, _xlsx_bytes(workbook), "typed.xlsx")

    result = parser.parse(artifact, no_network=True)

    assert [(table.name, table.visibility) for table in result.tables] == [
        ("Visible", "visible"),
        ("Hidden", "hidden"),
        ("Very Hidden", "veryHidden"),
    ]
    visible_table = result.tables[0]
    assert [cell.value for cell in visible_table.cells if cell.row == 1] == ["Merged heading"]
    assert next(cell for cell in visible_table.cells if cell.locator.value == "Visible!B2").value == (
        date(2024, 1, 1)
    )
    assert next(cell for cell in visible_table.cells if cell.locator.value == "Visible!B3").value is True
    margin = next(cell for cell in visible_table.cells if cell.locator.value == "Visible!B4")
    assert margin.value == Decimal("0.25")
    assert margin.unit == "percent"


def test_csv_bom_quotes_newlines_and_dangerous_prefixes_remain_inert_data(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    payload = (
        '\ufeffMetric,2025,Note\r\n'
        'ARR,"1000","line one\nline two"\r\n'
        'Formula,"=2+3","@cmd"\r\n'
    ).encode("utf-8")
    artifact = _persist(store, payload, "plan.csv", media_type="text/csv")

    table = parser.parse(artifact, no_network=True).tables[0]

    assert table.name == "plan"
    assert next(cell for cell in table.cells if cell.locator.value == "R2C3").value == (
        "line one\nline two"
    )
    injected = next(cell for cell in table.cells if cell.locator.value == "R3C2")
    assert injected.value == "=2+3"
    assert injected.locator.kind == "csv_cell"
    assert next(cell for cell in table.cells if cell.locator.value == "R3C3").value == "@cmd"


def test_csv_allowlisted_cp1252_decoding_is_deterministic(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    artifact = _persist(
        store,
        "Metric,Note\r\nARR,caf\u00e9\r\n".encode("cp1252"),
        "latin.csv",
        media_type="text/csv",
    )

    result = parser.parse(artifact, no_network=True)

    assert result.encoding == "cp1252"
    assert next(cell for cell in result.tables[0].cells if cell.locator.value == "R2C2").value == (
        "caf\u00e9"
    )


@pytest.mark.parametrize(
    ("limits", "payload", "expected_code"),
    [
        (SpreadsheetLimits(max_rows=1), b"a,b\r\n1,2\r\n", "csv_row_limit_exceeded"),
        (SpreadsheetLimits(max_columns=1), b"a,b\r\n", "csv_column_limit_exceeded"),
        (SpreadsheetLimits(max_field_chars=3), b"field\r\n", "csv_field_limit_exceeded"),
        (SpreadsheetLimits(max_bytes=3), b"abcd", "spreadsheet_byte_limit_exceeded"),
    ],
)
def test_csv_configured_quotas_return_typed_outcomes_without_raw_data(
    tmp_path: Path,
    limits: SpreadsheetLimits,
    payload: bytes,
    expected_code: str,
) -> None:
    store = LocalArtifactStore(tmp_path / "artifact-store")
    normalizer = TableNormalizationService(
        artifact_store=store,
        database_path=tmp_path / "normalized.duckdb",
    )
    parser = SpreadsheetParser(artifact_store=store, normalizer=normalizer, limits=limits)
    artifact = _persist(store, payload, "oversized.csv", media_type="text/csv")

    result = parser.parse(artifact, no_network=True)

    assert result.status == "quota_exceeded"
    assert result.error_code == expected_code
    assert payload.decode("ascii", errors="ignore") not in repr(result)


def test_malformed_csv_is_typed_and_does_not_leak_input(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    payload = b'Metric,Value\r\nSECRET,"unterminated\r\n'
    artifact = _persist(store, payload, "malformed.csv", media_type="text/csv")

    result = parser.parse(artifact, no_network=True)

    assert result.status == "malformed"
    assert result.error_code == "malformed_csv"
    assert "SECRET" not in repr(result)


def test_normalized_snapshot_is_deterministic_stored_by_hash_and_repr_is_private(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    payload = b"Metric,2025\r\nSECRET-ARR,1000\r\n"
    artifact = _persist(store, payload, "private.csv", media_type="text/csv")

    first = parser.parse(artifact, no_network=True).tables[0]
    second = parser.parse(artifact, no_network=True).tables[0]

    assert first.snapshot_hash == second.snapshot_hash
    assert first.snapshot_ref == first.snapshot_hash
    stored = json.loads(store.read_bytes(first.snapshot_ref))
    assert stored["name"] == "private"
    assert stored["cells"][2]["value"] == "SECRET-ARR"
    assert "SECRET-ARR" not in repr(first)
    assert "SECRET-ARR" not in repr(first.cells[2])


def test_duckdb_snapshot_write_is_transactional_parameterized_and_identifier_safe(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, database_path = spreadsheet_runtime
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales'; DROP TABLE x;--"
    sheet["A1"] = "safe"
    artifact = _persist(store, _xlsx_bytes(workbook), "identifier.xlsx")

    parsed = parser.parse(artifact, no_network=True)

    with duckdb.connect(str(database_path), read_only=True) as connection:
        rows = connection.execute(
            "SELECT table_name, snapshot_hash FROM normalized_table_snapshots"
        ).fetchall()
    assert rows == [(sheet.title, parsed.tables[0].snapshot_hash)]


def test_only_period_and_unit_validated_numeric_cells_become_evidence_facts(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evidence"
    sheet["A1"] = "Metric"
    sheet["B1"] = "2025"
    sheet["C1"] = "Unknown period"
    sheet["A2"] = "Revenue"
    sheet["B2"] = 1250.50
    sheet["B2"].number_format = '$#,##0.00'
    sheet["C2"] = 500
    sheet["C2"].number_format = '$#,##0'
    sheet["A3"] = "Headcount"
    sheet["B3"] = 12
    artifact = _persist(store, _xlsx_bytes(workbook), "evidence.xlsx")

    result = parser.parse(artifact, no_network=True)

    assert [(fact.name, fact.value, fact.unit, fact.period) for fact in result.evidence_facts] == [
        ("revenue", Decimal("1250.5"), "USD", "2025")
    ]
    cells = {cell.locator.value: cell for cell in result.tables[0].cells}
    assert cells["Evidence!B2"].status == "candidate"
    assert cells["Evidence!C2"].status == "insufficient_data"
    assert cells["Evidence!B3"].status == "insufficient_data"


def test_startup_metric_input_labels_become_spreadsheet_evidence_facts(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    payload = b"Metric,USD\r\nAs Of,2026-12-31\r\nCOGS,336000\r\n"
    artifact = _persist(store, payload, "startup-metrics.csv", media_type="text/csv")

    result = parser.parse(artifact, no_network=True)

    assert [(fact.name, fact.value, fact.unit, fact.period) for fact in result.evidence_facts] == [
        ("cogs", Decimal("336000"), "USD", "2026")
    ]


def test_unknown_and_deferred_startup_labels_do_not_become_evidence_fact_names(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Labels"
    sheet["B1"] = "2025"
    sheet["A2"] = "Mystery Sales Number"
    sheet["B2"] = 100
    sheet["B2"].number_format = '$#,##0'
    sheet["A3"] = "ARR"
    sheet["B3"] = 200
    sheet["B3"].number_format = '$#,##0'
    artifact = _persist(store, _xlsx_bytes(workbook), "unknown-labels.xlsx")

    result = parser.parse(artifact, no_network=True)

    assert result.evidence_facts == []
    numeric_cells = [cell for cell in result.tables[0].cells if isinstance(cell.value, Decimal)]
    assert [cell.status for cell in numeric_cells] == ["insufficient_data", "insufficient_data"]


def test_market_as_of_fact_is_not_promoted_with_financial_period(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Market"
    sheet["B1"] = "2025"
    sheet["A2"] = "Market Cap"
    sheet["B2"] = 500
    sheet["B2"].number_format = '$#,##0'
    artifact = _persist(store, _xlsx_bytes(workbook), "market-cap.xlsx")

    result = parser.parse(artifact, no_network=True)
    numeric = next(cell for cell in result.tables[0].cells if cell.value == Decimal("500"))

    assert numeric.status == "insufficient_data"
    assert result.evidence_facts == []


@pytest.mark.parametrize(
    ("header", "expected_period", "expected_fact_count"),
    [
        ("Q3 2025", "2025-Q3", 1),
        ("2025-Q3", "2025-Q3", 1),
        ("FY2025", None, 0),
        ("2025-Q5", None, 0),
        ("2025-08-12", None, 0),
    ],
)
def test_periods_are_canonicalized_and_validated_against_metric_engine_contract(
    spreadsheet_runtime: tuple[SpreadsheetParser, LocalArtifactStore, Path],
    header: str,
    expected_period: str | None,
    expected_fact_count: int,
) -> None:
    parser, store, _database_path = spreadsheet_runtime
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Periods"
    sheet["B1"] = header
    sheet["A2"] = "Revenue"
    sheet["B2"] = 100
    sheet["B2"].number_format = '$#,##0'
    artifact = _persist(store, _xlsx_bytes(workbook), "periods.xlsx")

    result = parser.parse(artifact, no_network=True)
    numeric = next(cell for cell in result.tables[0].cells if cell.value == Decimal("100"))

    assert numeric.period == expected_period
    assert len(result.evidence_facts) == expected_fact_count
    if expected_fact_count:
        assert result.evidence_facts[0].period == expected_period
    else:
        assert numeric.status == "insufficient_data"


def _xlsx_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _with_external_link(payload: bytes) -> bytes:
    source = BytesIO(payload)
    output = BytesIO()
    with zipfile.ZipFile(source) as original, zipfile.ZipFile(output, "w") as rewritten:
        for item in original.infolist():
            rewritten.writestr(item, original.read(item.filename))
        rewritten.writestr("xl/externalLinks/externalLink1.xml", b"<externalLink/>")
    return output.getvalue()


def _with_zip_member(payload: bytes, member_name: str, member_payload: bytes) -> bytes:
    source = BytesIO(payload)
    output = BytesIO()
    with zipfile.ZipFile(source) as original, zipfile.ZipFile(
        output, "w", compression=zipfile.ZIP_DEFLATED
    ) as rewritten:
        for item in original.infolist():
            rewritten.writestr(item, original.read(item.filename))
        rewritten.writestr(member_name, member_payload)
    return output.getvalue()


def _persist(
    store: LocalArtifactStore,
    payload: bytes,
    source: str,
    *,
    media_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
) -> Artifact:
    digest = __import__("hashlib").sha256(payload).hexdigest()
    artifact = Artifact(
        id=uuid4(),
        case_id=CASE_ID,
        content_hash=digest,
        mime_type=media_type,
        source=source,
        retrieved_at=datetime.now(UTC),
        source_snapshot_hash=digest,
        sensitivity=SensitivityClass.RESTRICTED,
    )
    store.put_bytes(
        payload,
        media_type=media_type,
        artifact_id=artifact.id,
        source_snapshot_hash=digest,
        sensitivity=artifact.sensitivity,
    )
    return artifact
